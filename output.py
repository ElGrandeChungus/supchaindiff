import re
import logging
from copy import copy as _copy
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import Config

logger = logging.getLogger(__name__)

_PROHIBITED = re.compile(r"[\[\]:*?/\\]")

# Colour fills (8-char ARGB: FF + 6-char RGB)
FILL_ADDED = PatternFill("solid", fgColor="FFC6EFCE")
FILL_REMOVED = PatternFill("solid", fgColor="FFFFC7CE")
FILL_CHANGED_CURRENT = PatternFill("solid", fgColor="FFFFEB9C")
FILL_CHANGED_PRIOR = PatternFill("solid", fgColor="FFFCE4D6")
FILL_SEV_HIGH = PatternFill("solid", fgColor="FFFF0000")
FILL_SEV_MEDIUM = PatternFill("solid", fgColor="FFFFBF00")
FILL_SEV_LOW_NONE = PatternFill("solid", fgColor="FF00B050")
FILL_SEV_UNSCORED = PatternFill("solid", fgColor="FF808080")


def sanitize_ben_for_sheet(ben: str, prefix: str = "DIFF_") -> str:
    """Strip prohibited chars; truncate BEN portion to 26 chars; prepend prefix."""
    clean = _PROHIBITED.sub("", ben)
    max_ben = 31 - len(prefix)  # 26 when prefix="DIFF_"
    return f"{prefix}{clean[:max_ben]}"


def assign_sheet_names(bens: List[str]) -> Dict[str, str]:
    """Return {original_ben: sheet_name} with collision resolution."""
    seen: Dict[str, int] = {}
    result: Dict[str, str] = {}
    for ben in bens:
        base = sanitize_ben_for_sheet(ben)
        if base not in seen:
            seen[base] = 1
            result[ben] = base
        else:
            seen[base] += 1
            n = seen[base]
            suffix = f"_{n}"
            result[ben] = base[: 31 - len(suffix)] + suffix
    return result


def write_diff_index(
    ws,
    groups: Dict[str, Dict[str, Any]],
    severities: Dict[str, Dict[str, Any]],
    sheet_names: Dict[str, str],
    skipped_bens: List[str],
    col_asymmetry: List[str],
    config: Config,
) -> None:
    """Write the DIFF_INDEX sheet to the given worksheet."""
    aliases = ", ".join(config.ben_aliases)
    ben_header = f"BEN (also known as: {aliases})" if aliases else "BEN"

    headers = [
        ben_header, "Sheet Name", "Added", "Removed", "Changed", "Unchanged",
        "Status", "Severity", "LLM Note",
    ]
    for col_i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_i, value=h)
        cell.font = Font(bold=True)

    row = 2
    for ben, group in groups.items():
        sev_info = severities.get(ben, {"severity": "UNSCORED", "note": ""})
        severity = sev_info.get("severity", "UNSCORED")
        ws.cell(row=row, column=1, value=ben)
        ws.cell(row=row, column=2, value=sheet_names.get(ben, ""))
        ws.cell(row=row, column=3, value=len(group["added"]))
        ws.cell(row=row, column=4, value=len(group["removed"]))
        ws.cell(row=row, column=5, value=len(group["changed_b"]))
        ws.cell(row=row, column=6, value=len(group["unchanged"]))
        ws.cell(row=row, column=7, value="OK")
        sev_cell = ws.cell(row=row, column=8, value=severity)
        sev_cell.fill = _severity_fill(severity)
        ws.cell(row=row, column=9, value=sev_info.get("note", ""))
        row += 1

    for ben in skipped_bens:
        ws.cell(row=row, column=1, value=ben)
        ws.cell(row=row, column=7, value="SKIPPED — DUPLICATE KEYS")
        row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"

    if col_asymmetry:
        row += 1
        note = f"Note: Column asymmetry detected — columns present in only one sheet: {', '.join(col_asymmetry)}"
        ws.cell(row=row, column=1, value=note)


def _severity_fill(severity: str) -> PatternFill:
    return {
        "HIGH": FILL_SEV_HIGH,
        "MEDIUM": FILL_SEV_MEDIUM,
        "LOW": FILL_SEV_LOW_NONE,
        "NONE": FILL_SEV_LOW_NONE,
    }.get(severity, FILL_SEV_UNSCORED)


def write_tool_sheet(
    ws,
    group: Dict[str, Any],
    non_key_cols: List[str],
    config: Config,
) -> None:
    """Write one per-tool diff sheet with four sections: ADDED, REMOVED, CHANGED, UNCHANGED."""
    keys = config.composite_key_columns
    all_cols = keys + non_key_cols
    prior_headers = [f"{c}_prior" for c in non_key_cols]

    # Build lookup: (ben_str, part_str) -> set of changed field names
    changed_lookup: Dict[Tuple, set] = {}
    for fd in group["field_diffs"]:
        k = (str(fd["ben"]), str(fd["part"]))
        changed_lookup.setdefault(k, set()).add(fd["field"])

    current_row = 1

    # Section 1: ADDED
    current_row = _write_section_header(ws, current_row, "ADDED")
    current_row = _write_col_headers(ws, current_row, all_cols)
    for _, data_row in group["added"].iterrows():
        for col_i, col in enumerate(all_cols, 1):
            cell = ws.cell(row=current_row, column=col_i, value=_safe_val(data_row.get(col)))
            cell.fill = FILL_ADDED
        ws.row_dimensions[current_row].outline_level = 0
        current_row += 1

    # Section 2: REMOVED
    current_row = _write_section_header(ws, current_row, "REMOVED")
    current_row = _write_col_headers(ws, current_row, all_cols)
    for _, data_row in group["removed"].iterrows():
        for col_i, col in enumerate(all_cols, 1):
            cell = ws.cell(row=current_row, column=col_i, value=_safe_val(data_row.get(col)))
            cell.fill = FILL_REMOVED
        ws.row_dimensions[current_row].outline_level = 0
        current_row += 1

    # Section 3: CHANGED
    full_changed_headers = all_cols + prior_headers
    current_row = _write_section_header(ws, current_row, "CHANGED")
    current_row = _write_col_headers(ws, current_row, full_changed_headers)
    changed_b = group["changed_b"]
    changed_a = group["changed_a"]
    for i in range(len(changed_b)):
        row_b = changed_b.iloc[i]
        row_a = changed_a.iloc[i] if not changed_a.empty else None
        ben_val = str(row_b.get(config.ben_column, ""))
        part_val = str(row_b.get(config.part_number_column, ""))
        row_changed_fields = changed_lookup.get((ben_val, part_val), set())

        # Group 1: current values (Sheet B)
        for col_i, col in enumerate(all_cols, 1):
            cell = ws.cell(row=current_row, column=col_i, value=_safe_val(row_b.get(col)))
            if col in row_changed_fields:
                cell.fill = FILL_CHANGED_CURRENT

        # Group 2: prior values (_prior columns)
        for j, col in enumerate(non_key_cols):
            col_i = len(all_cols) + j + 1
            if col in row_changed_fields and row_a is not None:
                prior_val = _safe_val(row_a.get(col))
                cell = ws.cell(row=current_row, column=col_i, value=prior_val)
                cell.fill = FILL_CHANGED_PRIOR
            else:
                ws.cell(row=current_row, column=col_i, value="")

        ws.row_dimensions[current_row].outline_level = 0
        current_row += 1

    # Section 4: UNCHANGED (collapsed by default)
    current_row = _write_section_header(ws, current_row, "UNCHANGED")
    current_row = _write_col_headers(ws, current_row, all_cols)
    for _, data_row in group["unchanged"].iterrows():
        for col_i, col in enumerate(all_cols, 1):
            ws.cell(row=current_row, column=col_i, value=_safe_val(data_row.get(col)))
        dim = ws.row_dimensions[current_row]
        dim.outline_level = 1
        dim.hidden = True
        current_row += 1


def _write_section_header(ws, row: int, label: str) -> int:
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = Font(bold=True)
    ws.row_dimensions[row].outline_level = 0
    return row + 1


def _write_col_headers(ws, row: int, headers: List[str]) -> int:
    for col_i, h in enumerate(headers, 1):
        ws.cell(row=row, column=col_i, value=h)
    ws.row_dimensions[row].outline_level = 0
    return row + 1


def _safe_val(v: Any) -> Any:
    """Convert NaN/NaT to None for openpyxl cell writing."""
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    return v


def write_exec_summary(
    ws, summary_text: str, config: Config, date_a: str, date_b: str
) -> None:
    """Write the EXEC_SUMMARY sheet."""
    ws.merge_cells("A1:H1")
    cell = ws.cell(row=1, column=1, value=summary_text)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[1].height = 200

    metadata_start = 3
    ws.cell(row=metadata_start, column=1, value="Model:").font = Font(bold=True)
    ws.cell(row=metadata_start, column=2, value=config.llm_model)
    ws.cell(row=metadata_start + 1, column=1, value="Generated:").font = Font(bold=True)
    ws.cell(row=metadata_start + 1, column=2, value=datetime.utcnow().isoformat() + "Z")
    ws.cell(row=metadata_start + 2, column=1, value="Disclaimer:").font = Font(bold=True)
    ws.cell(
        row=metadata_start + 2,
        column=2,
        value=(
            "This summary was generated by an AI model from deterministic diff data. "
            "All values originate from the source spreadsheet."
        ),
    )


def copy_sheet(source_ws, target_wb, title: str):
    """Copy sheet values and styles from source to a new sheet in target_wb."""
    target_ws = target_wb.create_sheet(title=title)
    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = _copy(cell.font)
                new_cell.border = _copy(cell.border)
                new_cell.fill = _copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = _copy(cell.protection)
                new_cell.alignment = _copy(cell.alignment)
    for key, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[key].width = dim.width
    return target_ws


def write_output_file(
    input_filepath: str,
    output_filepath: str,
    groups: Dict[str, Dict[str, Any]],
    severities: Dict[str, Dict[str, Any]],
    skipped_bens: List[str],
    col_asymmetry: List[str],
    exec_summary: str,
    date_a: str,
    date_b: str,
    config: Config,
    non_key_cols: List[str],
) -> None:
    """Write the complete output xlsx.

    Sheet order: all original sheets (copied verbatim) -> DIFF_INDEX ->
    one DIFF_{BEN} per tool -> EXEC_SUMMARY_{dateA}_{dateB}.
    The input file is never modified.
    """
    input_wb = load_workbook(input_filepath, read_only=False, data_only=True)
    output_wb = Workbook()
    output_wb.remove(output_wb.active)  # remove default empty sheet

    # 1. Copy all original sheets verbatim
    for sheet_name in input_wb.sheetnames:
        copy_sheet(input_wb[sheet_name], output_wb, sheet_name)

    # 2. DIFF_INDEX
    diff_index_ws = output_wb.create_sheet("DIFF_INDEX")
    sheet_name_map = assign_sheet_names(list(groups.keys()))
    write_diff_index(
        diff_index_ws, groups, severities, sheet_name_map,
        skipped_bens, col_asymmetry, config,
    )

    # 3. Per-tool diff sheets
    for ben, group in groups.items():
        sheet_name = sheet_name_map[ben]
        tool_ws = output_wb.create_sheet(sheet_name)
        write_tool_sheet(tool_ws, group, non_key_cols, config)

    # 4. EXEC_SUMMARY sheet
    exec_ws = output_wb.create_sheet(f"EXEC_SUMMARY_{date_a}_{date_b}")
    write_exec_summary(exec_ws, exec_summary, config, date_a, date_b)

    output_wb.save(output_filepath)
    logger.info("Output written to %s", output_filepath)
