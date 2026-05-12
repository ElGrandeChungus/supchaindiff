import pytest
from output import sanitize_ben_for_sheet, assign_sheet_names


def test_sanitize_strips_prohibited_chars():
    assert sanitize_ben_for_sheet("ABC[123]") == "DIFF_ABC123"
    assert sanitize_ben_for_sheet("A:B/C") == "DIFF_ABC"
    assert sanitize_ben_for_sheet("X*Y?Z") == "DIFF_XYZ"
    assert sanitize_ben_for_sheet(r"A\B") == "DIFF_AB"


def test_sanitize_truncates_ben_to_26_chars():
    long_ben = "A" * 50
    name = sanitize_ben_for_sheet(long_ben)
    assert len(name) == 31  # DIFF_ (5) + 26
    assert name.startswith("DIFF_")


def test_sanitize_short_ben_not_truncated():
    name = sanitize_ben_for_sheet("T1")
    assert name == "DIFF_T1"
    assert len(name) <= 31


def test_sanitize_result_always_within_31_chars():
    for ben in ["A" * 30, "B:C[D]" * 10, "X" * 26]:
        name = sanitize_ben_for_sheet(ben)
        assert len(name) <= 31, f"Name too long: {name!r}"


def test_assign_sheet_names_no_collision():
    bens = ["TOOL_A", "TOOL_B", "TOOL_C"]
    mapping = assign_sheet_names(bens)
    sheet_names = list(mapping.values())
    assert len(set(sheet_names)) == 3
    assert all(n.startswith("DIFF_") for n in sheet_names)


def test_assign_sheet_names_resolves_collision():
    # Two BENs that differ only after 26 chars will collide after truncation
    ben1 = "A" * 26 + "EXTRA1"
    ben2 = "A" * 26 + "EXTRA2"
    mapping = assign_sheet_names([ben1, ben2])
    sheet_names = list(mapping.values())
    assert len(set(sheet_names)) == 2
    assert all(len(n) <= 31 for n in sheet_names)


def test_assign_sheet_names_maps_original_ben():
    bens = ["MY_TOOL"]
    mapping = assign_sheet_names(bens)
    assert "MY_TOOL" in mapping
    assert mapping["MY_TOOL"] == "DIFF_MY_TOOL"


import pandas as pd
from openpyxl import Workbook
from output import write_diff_index
from config import Config


def _empty_df(config):
    return pd.DataFrame(columns=[config.ben_column, config.part_number_column])


def _make_severity_result(tool_id, severity="NONE", note=""):
    return {"tool_id": tool_id, "severity": severity, "note": note}


def test_diff_index_has_one_row_per_ben():
    config = Config()
    wb = Workbook()
    ws = wb.active
    ws.title = "DIFF_INDEX"

    groups = {
        "T1": {"added": _empty_df(config), "removed": _empty_df(config),
               "changed_b": _empty_df(config), "unchanged": _empty_df(config), "field_diffs": []},
        "T2": {"added": _empty_df(config), "removed": _empty_df(config),
               "changed_b": _empty_df(config), "unchanged": _empty_df(config), "field_diffs": []},
    }
    severities = {"T1": _make_severity_result("T1"), "T2": _make_severity_result("T2")}
    sheet_names = {"T1": "DIFF_T1", "T2": "DIFF_T2"}

    write_diff_index(ws, groups, severities, sheet_names, skipped_bens=[], col_asymmetry=[], config=config)

    # Row 1 = header, rows 2+ = data
    data_vals = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    data_vals = [v for v in data_vals if v is not None]
    assert len(data_vals) == 2


def test_diff_index_header_uses_ben_alias_annotation():
    config = Config()
    wb = Workbook()
    ws = wb.active

    write_diff_index(ws, {}, {}, {}, skipped_bens=[], col_asymmetry=[], config=config)

    header = ws.cell(row=1, column=1).value
    assert "BEN" in header
    assert "P1A" in header
    assert "FCID" in header


def test_diff_index_skipped_ben_shows_status():
    config = Config()
    wb = Workbook()
    ws = wb.active

    write_diff_index(ws, {}, {}, {}, skipped_bens=["BAD_TOOL"], col_asymmetry=[], config=config)

    statuses = [ws.cell(row=r, column=7).value for r in range(2, ws.max_row + 1)]
    assert any("SKIPPED" in str(s) for s in statuses if s)


def test_diff_index_ok_rows_have_ok_status():
    config = Config()
    wb = Workbook()
    ws = wb.active

    groups = {
        "T1": {"added": _empty_df(config), "removed": _empty_df(config),
               "changed_b": _empty_df(config), "unchanged": _empty_df(config), "field_diffs": []},
    }
    severities = {"T1": _make_severity_result("T1", severity="HIGH")}
    sheet_names = {"T1": "DIFF_T1"}

    write_diff_index(ws, groups, severities, sheet_names, skipped_bens=[], col_asymmetry=[], config=config)

    status_val = ws.cell(row=2, column=7).value
    assert status_val == "OK"


def test_diff_index_col_asymmetry_note_added():
    config = Config()
    wb = Workbook()
    ws = wb.active

    write_diff_index(ws, {}, {}, {}, skipped_bens=[], col_asymmetry=["WeirdCol"], config=config)

    all_values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert any("WeirdCol" in str(v) for v in all_values if v)


from output import write_tool_sheet


def test_tool_sheet_has_four_section_headers():
    config = Config()
    wb = Workbook()
    ws = wb.create_sheet("DIFF_T1")
    group = {
        "added": pd.DataFrame({"BEN": ["T1"], "Part Number": ["P_NEW"], "Qty": [5]}),
        "removed": pd.DataFrame({"BEN": ["T1"], "Part Number": ["P_OLD"], "Qty": [3]}),
        "changed_b": pd.DataFrame({"BEN": ["T1"], "Part Number": ["P1"], "Qty": [10]}),
        "changed_a": pd.DataFrame({"BEN": ["T1"], "Part Number": ["P1"], "Qty": [5]}),
        "unchanged": pd.DataFrame({"BEN": ["T1"], "Part Number": ["P2"], "Qty": [2]}),
        "field_diffs": [{"ben": "T1", "part": "P1", "field": "Qty", "old": 5, "new": 10}],
    }
    non_key = ["Qty"]
    write_tool_sheet(ws, group, non_key, config)
    all_values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert "ADDED" in all_values
    assert "REMOVED" in all_values
    assert "CHANGED" in all_values
    assert "UNCHANGED" in all_values


def test_tool_sheet_unchanged_rows_have_outline_level_1():
    config = Config()
    wb = Workbook()
    ws = wb.create_sheet("DIFF_T1")
    group = {
        "added": pd.DataFrame(columns=["BEN", "Part Number", "Qty"]),
        "removed": pd.DataFrame(columns=["BEN", "Part Number", "Qty"]),
        "changed_b": pd.DataFrame(columns=["BEN", "Part Number", "Qty"]),
        "changed_a": pd.DataFrame(columns=["BEN", "Part Number", "Qty"]),
        "unchanged": pd.DataFrame({"BEN": ["T1"], "Part Number": ["P1"], "Qty": [5]}),
        "field_diffs": [],
    }
    write_tool_sheet(ws, group, ["Qty"], config)
    hidden_rows = [r for r in range(1, ws.max_row + 2) if ws.row_dimensions[r].outline_level == 1]
    assert len(hidden_rows) == 1


def test_tool_sheet_changed_section_has_prior_columns():
    config = Config()
    wb = Workbook()
    ws = wb.create_sheet("DIFF_T1")
    group = {
        "added": pd.DataFrame(columns=["BEN", "Part Number", "Qty"]),
        "removed": pd.DataFrame(columns=["BEN", "Part Number", "Qty"]),
        "changed_b": pd.DataFrame({"BEN": ["T1"], "Part Number": ["P1"], "Qty": [10]}),
        "changed_a": pd.DataFrame({"BEN": ["T1"], "Part Number": ["P1"], "Qty": [5]}),
        "unchanged": pd.DataFrame(columns=["BEN", "Part Number", "Qty"]),
        "field_diffs": [{"ben": "T1", "part": "P1", "field": "Qty", "old": 5, "new": 10}],
    }
    write_tool_sheet(ws, group, ["Qty"], config)
    all_cell_values = set()
    for row in ws.iter_rows():
        for cell in row:
            all_cell_values.add(cell.value)
    assert "Qty_prior" in all_cell_values


def test_tool_sheet_added_rows_have_green_fill():
    config = Config()
    wb = Workbook()
    ws = wb.create_sheet("DIFF_T1")
    group = {
        "added": pd.DataFrame({"BEN": ["T1"], "Part Number": ["P_NEW"], "Qty": [5]}),
        "removed": pd.DataFrame(columns=["BEN", "Part Number", "Qty"]),
        "changed_b": pd.DataFrame(columns=["BEN", "Part Number", "Qty"]),
        "changed_a": pd.DataFrame(columns=["BEN", "Part Number", "Qty"]),
        "unchanged": pd.DataFrame(columns=["BEN", "Part Number", "Qty"]),
        "field_diffs": [],
    }
    write_tool_sheet(ws, group, ["Qty"], config)
    # Find cells with green fill (ADDED fill color)
    green_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb == "FFC6EFCE":
                green_cells.append(cell)
    assert len(green_cells) > 0
